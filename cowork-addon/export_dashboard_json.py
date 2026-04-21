#!/usr/bin/env python3
"""
export_dashboard_json.py

Runs inside Cowork after `VOTES Metrics.xlsx` is saved. Reads the workbook,
assembles a data.json with per-metric time series, writes it to the same
project folder.

Prefers the 'raw data' tab (single-sheet layout with one row per week).
Falls back to iterating date-named tabs if 'raw data' is missing.

The dashboard repo's GitHub Action consumes this file. If this script fails,
the xlsx remains the source of truth and the Action's fallback path will
parse the xlsx directly.

Contract (do not change without also updating the dashboard repo):

  {
    "generated_at": "<ISO 8601 UTC timestamp>",
    "weeks": ["YYYY-MM-DD", ...],                       # sorted ascending
    "series": {
      "V": [<number|null>, ...],                        # aligned to weeks
      "O": [<number|null>, ...],
      "T": [<number|null>, ...],
      "E": [<number|null>, ...],
      "S": [<number|null>, ...]
    },
    "metric_labels_from_sheet": {
      "V": "V - Velocity",
      ...
    }
  }
"""
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

WORKBOOK_PATH = Path("VOTES Metrics.xlsx")
OUTPUT_PATH = Path("data.json")
REQUIRED_IDS = {"V", "O", "T", "E", "S"}
RAW_DATA_SHEET_NAME = "raw data"
DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def coerce_number(raw):
    """Return int/float for numeric-looking values, None otherwise."""
    if isinstance(raw, (int, float)):
        return int(raw) if float(raw).is_integer() else float(raw)
    try:
        val = float(raw)
    except (TypeError, ValueError):
        return None
    return int(val) if val.is_integer() else val


def extract_metric_id(label):
    """'V - Velocity' -> 'V', else None."""
    m = re.match(r"^([A-Z])\s*-\s*", str(label or ""))
    return m.group(1) if m else None


def find_raw_data_sheet(wb):
    """Case-insensitive match for 'raw data' tab, or None."""
    for name in wb.sheetnames:
        if name.strip().lower() == RAW_DATA_SHEET_NAME:
            return wb[name]
    return None


def parse_raw_data(ws):
    """
    Parse the 'raw data' sheet.
    Expected layout: Date | V - ... | O - ... | T - ... | E - ... | S - ...
    Returns (per_week, labels) or raises ValueError.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("'raw data' sheet is empty")

    header = rows[0]
    col_to_mid = {}
    labels = {}
    for idx, cell in enumerate(header):
        if idx == 0:
            continue  # date column
        mid = extract_metric_id(cell)
        if mid is None:
            print(f"[export_dashboard_json] 'raw data' header col {idx} {cell!r} has no 'X - ' prefix; ignoring")
            continue
        col_to_mid[idx] = mid
        labels[mid] = str(cell)

    if not col_to_mid:
        raise ValueError("'raw data' sheet has no recognized metric columns")

    per_week = {}
    for r in rows[1:]:
        if not r or r[0] is None:
            continue
        date_str = str(r[0]).strip()
        if not DATE_RE.fullmatch(date_str):
            print(f"[export_dashboard_json] 'raw data' row with non-ISO date {date_str!r}; skipping")
            continue
        week_values = {}
        for col_idx, mid in col_to_mid.items():
            week_values[mid] = coerce_number(r[col_idx]) if col_idx < len(r) else None
        per_week[date_str] = week_values

    if not per_week:
        raise ValueError("'raw data' sheet has no data rows")

    return per_week, labels


def parse_date_tabs(wb):
    """Fallback: parse every YYYY-MM-DD tab, using column A=metric, C=current."""
    per_week = {}
    labels = {}
    for name in wb.sheetnames:
        if not DATE_RE.fullmatch(name):
            print(f"[export_dashboard_json] skipping non-date sheet: {name!r}")
            continue
        ws = wb[name]
        week_values = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 3:
                continue
            metric_name = row[0]
            current = row[2]
            if not metric_name:
                continue
            mid = extract_metric_id(metric_name)
            if mid is None:
                continue
            labels.setdefault(mid, str(metric_name))
            week_values[mid] = coerce_number(current)
        per_week[name] = week_values
    return per_week, labels


def main():
    if not WORKBOOK_PATH.exists():
        sys.exit(f"[export_dashboard_json] {WORKBOOK_PATH} not found in CWD")

    wb = load_workbook(WORKBOOK_PATH, data_only=True)

    # Prefer 'raw data' tab, fall back to date tabs
    raw_ws = find_raw_data_sheet(wb)
    if raw_ws is not None:
        print(f"[export_dashboard_json] using {raw_ws.title!r} tab (single-sheet layout).")
        try:
            per_week, labels = parse_raw_data(raw_ws)
        except ValueError as e:
            print(f"[export_dashboard_json] {raw_ws.title!r} tab unusable ({e}); falling back to date tabs.")
            per_week, labels = parse_date_tabs(wb)
    else:
        print("[export_dashboard_json] no 'raw data' tab; using date tabs (fallback layout).")
        per_week, labels = parse_date_tabs(wb)

    if not per_week:
        sys.exit("[export_dashboard_json] no data found in workbook")

    weeks_sorted = sorted(per_week.keys())

    # Self-check: every week must have every required metric ID.
    # Aborting here prevents overwriting Drive's data.json with incomplete data;
    # the dashboard Action will fall back to parsing the xlsx directly.
    for week, mvals in per_week.items():
        missing = REQUIRED_IDS - set(mvals.keys())
        if missing:
            sys.exit(
                f"[export_dashboard_json] week {week} is missing metric rows: "
                f"{sorted(missing)}. Aborting without overwriting data.json."
            )

    all_ids = sorted(REQUIRED_IDS | set(labels.keys()))
    series = {
        mid: [per_week[w].get(mid) for w in weeks_sorted]
        for mid in all_ids
    }

    payload = {
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "weeks": weeks_sorted,
        "series": series,
        "metric_labels_from_sheet": labels,
    }

    OUTPUT_PATH.write_text(json.dumps(payload, indent=2) + "\n")
    print(
        f"[export_dashboard_json] wrote {OUTPUT_PATH} "
        f"({len(weeks_sorted)} weeks, {len(all_ids)} metrics)."
    )


if __name__ == "__main__":
    main()
